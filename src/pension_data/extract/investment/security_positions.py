"""Security-level holdings extraction and CAFR coverage reconciliation."""

from __future__ import annotations

import csv
import math
from collections import defaultdict
from dataclasses import dataclass
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any

from defusedxml import ElementTree  # type: ignore[import-untyped]

from pension_data.db.models.investment_positions import (
    HoldingsCoverageReport,
    PlanSecurityPosition,
    SecurityDisclosureState,
    SecurityPositionSource,
)
from pension_data.finite_guards import finite_or_none

_TOTAL_PLAN_COVERAGE_THRESHOLD = 0.95


@dataclass(frozen=True, slots=True)
class SecurityPositionInput:
    """Raw security-level holding from a public source."""

    security_name: str | None
    cusip: str | None
    ticker: str | None
    shares: float | None
    market_value_usd: float | None
    asset_class: str
    source: SecurityPositionSource
    as_of: str
    provenance_ref: str
    disclosure_state: SecurityDisclosureState = "disclosed"
    manager_name: str | None = None
    fund_name: str | None = None
    confidence: float = 1.0
    valid_from: str | None = None
    valid_to: str | None = None
    asserted_at: str | None = None
    amendment_accession: str | None = None


@dataclass(frozen=True, slots=True)
class AcfrAllocationInput:
    """ACFR total-plan anchor row used to label holdings coverage."""

    asset_class: str
    market_value_usd: float
    provenance_ref: str


def _text(element: ElementTree.Element, tag_name: str) -> str | None:
    for child in element.iter():
        if child.tag.rsplit("}", 1)[-1] == tag_name and child.text:
            value = child.text.strip()
            if value:
                return str(value)
    return None


def _float_or_none(value: str | None) -> float | None:
    if value is None:
        return None
    cleaned = value.strip().replace(",", "")
    if not cleaned:
        return None
    try:
        parsed = float(cleaned)
    except ValueError:
        return None
    if not math.isfinite(parsed):
        return None
    return parsed


def _normalize_cusip(value: str | None) -> str | None:
    if value is None:
        return None
    normalized = "".join(ch for ch in value.upper() if ch.isalnum())
    return normalized or None


def _security_id(*, cusip: str | None, ticker: str | None, security_name: str | None) -> str:
    if cusip:
        return f"cusip:{cusip}"
    if ticker:
        return f"ticker:{ticker.strip().upper()}"
    if security_name:
        return "name:" + " ".join(security_name.lower().split())
    msg = "security position requires cusip, ticker, or security_name"
    raise ValueError(msg)


def parse_13f_information_table_xml(
    xml_text: str,
    *,
    as_of: str,
    provenance_ref: str,
    asset_class: str = "public_equity",
) -> list[SecurityPositionInput]:
    """Parse an EDGAR 13F information table XML payload.

    13F values are reported in thousands of dollars, so `market_value_usd`
    multiplies the XML `value` field by 1,000.
    """
    root = ElementTree.fromstring(xml_text)
    rows: list[SecurityPositionInput] = []
    for info_table in root.iter():
        if info_table.tag.rsplit("}", 1)[-1] != "infoTable":
            continue
        security_name = _text(info_table, "nameOfIssuer")
        cusip = _normalize_cusip(_text(info_table, "cusip"))
        value_thousands = _float_or_none(_text(info_table, "value"))
        shares = _float_or_none(_text(info_table, "sshPrnamt"))
        rows.append(
            SecurityPositionInput(
                security_name=security_name,
                cusip=cusip,
                ticker=None,
                shares=shares,
                market_value_usd=(
                    round(value_thousands * 1000.0, 6) if value_thousands is not None else None
                ),
                asset_class=asset_class,
                source="13f",
                as_of=as_of,
                provenance_ref=provenance_ref,
            )
        )
    return rows


def _own_holdings_input_from_row(
    row: dict[str, str | None],
    *,
    as_of: str,
    provenance_ref: str,
    default_asset_class: str,
) -> SecurityPositionInput:
    """Build one own-holdings input from a header-keyed row (CSV or spreadsheet)."""
    security_name = row.get("security_name") or row.get("name")
    ticker = row.get("ticker")
    return SecurityPositionInput(
        security_name=security_name.strip() if security_name else None,
        cusip=_normalize_cusip(row.get("cusip")),
        ticker=ticker.strip().upper() if ticker and ticker.strip() else None,
        shares=_float_or_none(row.get("shares")),
        market_value_usd=_float_or_none(row.get("market_value_usd")),
        asset_class=(row.get("asset_class") or default_asset_class).strip().lower(),
        source="own_holdings_file",
        as_of=as_of,
        provenance_ref=provenance_ref,
        manager_name=(row.get("manager_name") or None),
        fund_name=(row.get("fund_name") or None),
    )


def load_own_holdings_csv(
    csv_text: str,
    *,
    as_of: str,
    provenance_ref: str,
    default_asset_class: str = "unknown",
) -> list[SecurityPositionInput]:
    """Load a public own-holdings CSV export into normalized inputs."""
    reader = csv.DictReader(StringIO(csv_text))
    return [
        _own_holdings_input_from_row(
            dict(row),
            as_of=as_of,
            provenance_ref=provenance_ref,
            default_asset_class=default_asset_class,
        )
        for row in reader
    ]


def load_own_holdings_xls(
    source: str | Path | bytes,
    *,
    as_of: str,
    provenance_ref: str,
    default_asset_class: str = "unknown",
    sheet_name: str | None = None,
) -> list[SecurityPositionInput]:
    """Load a public own-holdings ``.xlsx``/``.xls`` export into normalized inputs.

    ``source`` may be a filesystem path or the raw workbook bytes. The first row
    is treated as the header; header cells are matched case-insensitively against
    the same column names as :func:`load_own_holdings_csv`
    (``security_name``/``name``, ``cusip``, ``ticker``, ``shares``,
    ``market_value_usd``, ``asset_class``, ``manager_name``, ``fund_name``).

    ``openpyxl`` is an optional dependency (the ``source_collection`` extra); it is
    imported lazily so the base install and CI test gate stay dependency-light.
    """
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:  # pragma: no cover - exercised via importorskip in tests
        raise ModuleNotFoundError(
            "load_own_holdings_xls requires openpyxl; install the 'source_collection' extra"
        ) from exc

    workbook_source: Any = BytesIO(source) if isinstance(source, bytes) else source
    workbook = load_workbook(filename=workbook_source, read_only=True, data_only=True)
    try:
        worksheet = workbook[sheet_name] if sheet_name is not None else workbook.active
        row_iter = worksheet.iter_rows(values_only=True)
        try:
            header_cells = next(row_iter)
        except StopIteration:
            return []
        headers = [str(cell).strip().lower() if cell is not None else "" for cell in header_cells]
        inputs: list[SecurityPositionInput] = []
        for raw_row in row_iter:
            if all(cell is None for cell in raw_row):
                continue
            row: dict[str, str | None] = {}
            for header, cell in zip(headers, raw_row, strict=False):
                if not header:
                    continue
                row[header] = None if cell is None else str(cell)
            inputs.append(
                _own_holdings_input_from_row(
                    row,
                    as_of=as_of,
                    provenance_ref=provenance_ref,
                    default_asset_class=default_asset_class,
                )
            )
        return inputs
    finally:
        workbook.close()


def build_security_positions(
    *,
    plan_id: str,
    plan_period: str,
    rows: list[SecurityPositionInput],
) -> list[PlanSecurityPosition]:
    """Stage security-level positions with deterministic IDs and ordering."""
    positions = [
        PlanSecurityPosition(
            plan_id=plan_id,
            plan_period=plan_period,
            security_id=_security_id(
                cusip=row.cusip,
                ticker=row.ticker,
                security_name=row.security_name,
            ),
            security_name=row.security_name.strip() if row.security_name else None,
            cusip=row.cusip,
            ticker=row.ticker.strip().upper() if row.ticker else None,
            shares=row.shares,
            market_value_usd=row.market_value_usd,
            asset_class=row.asset_class.strip().lower(),
            source=row.source,
            as_of=row.as_of,
            disclosure_state=row.disclosure_state,
            provenance_ref=row.provenance_ref,
            manager_name=row.manager_name.strip() if row.manager_name else None,
            fund_name=row.fund_name.strip() if row.fund_name else None,
            confidence=max(0.0, min(1.0, row.confidence)),
            valid_from=row.valid_from or row.as_of,
            valid_to=row.valid_to,
            asserted_at=row.asserted_at or row.as_of,
            amendment_accession=row.amendment_accession,
        )
        for row in rows
    ]
    return sorted(
        positions,
        key=lambda row: (
            row.plan_id,
            row.plan_period,
            row.asset_class,
            row.security_id,
            row.provenance_ref,
        ),
    )


def reconcile_holdings_to_acfr(
    *,
    plan_id: str,
    plan_period: str,
    positions: list[PlanSecurityPosition],
    total_plan_assets_usd: float,
    acfr_allocations: list[AcfrAllocationInput],
) -> HoldingsCoverageReport:
    """Compute holdings coverage against ACFR total-plan asset values."""
    if total_plan_assets_usd <= 0.0 or not math.isfinite(total_plan_assets_usd):
        msg = "total_plan_assets_usd must be a positive finite value"
        raise ValueError(msg)

    collected_by_asset_class: dict[str, float] = defaultdict(float)
    scoped_provenance_refs: list[str] = []
    for position in positions:
        if (
            position.plan_id != plan_id
            or position.plan_period != plan_period
            or position.disclosure_state != "disclosed"
            or position.market_value_usd is None
        ):
            continue
        collected_by_asset_class[position.asset_class] += position.market_value_usd
        if position.provenance_ref:
            scoped_provenance_refs.append(position.provenance_ref)

    acfr_by_asset_class = {
        row.asset_class.strip().lower(): row.market_value_usd for row in acfr_allocations
    }
    combined_classes = sorted(set(collected_by_asset_class) | set(acfr_by_asset_class))
    collected_total = round(sum(collected_by_asset_class.values()), 6)
    provenance_refs = tuple(
        dict.fromkeys(
            [
                *scoped_provenance_refs,
                *(row.provenance_ref for row in acfr_allocations if row.provenance_ref),
            ]
        )
    )
    coverage_ratio = round(collected_total / total_plan_assets_usd, 6)
    scope_label = (
        "total-plan" if coverage_ratio >= _TOTAL_PLAN_COVERAGE_THRESHOLD else "equity-sleeve"
    )

    return HoldingsCoverageReport(
        plan_id=plan_id,
        plan_period=plan_period,
        total_plan_assets_usd=round(total_plan_assets_usd, 6),
        collected_market_value_usd=collected_total,
        coverage_ratio=coverage_ratio,
        scope_label=scope_label,
        by_asset_class={
            asset_class: round(
                collected_by_asset_class.get(asset_class, 0.0) / acfr_by_asset_class[asset_class],
                6,
            )
            for asset_class in combined_classes
            if acfr_by_asset_class.get(asset_class, 0.0) > 0.0
        },
        provenance_refs=provenance_refs,
    )


@dataclass(frozen=True, slots=True)
class Ab2833AltDisclosureInput:
    """Raw AB 2833 alternative-investment fee/valuation disclosure row.

    California AB 2833 (Gov. Code §7514.7) requires public pension plans to
    publish, per alternative-investment vehicle, the plan's fair value plus the
    management fees and carried interest borne. This captures that disclosure so
    alts (which 13F never covers) can contribute to total-plan coverage.
    """

    fund_name: str
    asset_class: str
    reported_fair_value_usd: float | None
    management_fees_usd: float | None
    carried_interest_usd: float | None
    as_of: str
    provenance_ref: str
    manager_name: str | None = None
    explicit_not_disclosed: bool = False


@dataclass(frozen=True, slots=True)
class Ab2833AltDisclosure:
    """Captured AB 2833 alts disclosure with summed fees and a disclosure state."""

    fund_name: str
    manager_name: str | None
    asset_class: str
    reported_fair_value_usd: float | None
    management_fees_usd: float | None
    carried_interest_usd: float | None
    total_fees_usd: float | None
    disclosure_state: SecurityDisclosureState
    as_of: str
    provenance_ref: str


def capture_ab2833_alt_disclosures(
    rows: list[Ab2833AltDisclosureInput],
) -> list[Ab2833AltDisclosure]:
    """Normalize AB 2833 alts disclosures, summing fees under finite guards.

    Fees are summed only over finite components; a non-finite fee is dropped from
    the total rather than poisoning it to NaN. A row with an explicit
    non-disclosure flag or no fair value is marked ``not_disclosed``.
    """
    captured: list[Ab2833AltDisclosure] = []
    for row in rows:
        fair_value = finite_or_none(row.reported_fair_value_usd)
        management_fees = finite_or_none(row.management_fees_usd)
        carried_interest = finite_or_none(row.carried_interest_usd)
        fee_components = [fee for fee in (management_fees, carried_interest) if fee is not None]
        total_fees = round(sum(fee_components), 6) if fee_components else None
        disclosure_state: SecurityDisclosureState = (
            "not_disclosed" if row.explicit_not_disclosed or fair_value is None else "disclosed"
        )
        captured.append(
            Ab2833AltDisclosure(
                fund_name=row.fund_name.strip(),
                manager_name=row.manager_name.strip() if row.manager_name else None,
                asset_class=row.asset_class.strip().lower(),
                reported_fair_value_usd=fair_value,
                management_fees_usd=management_fees,
                carried_interest_usd=carried_interest,
                total_fees_usd=total_fees,
                disclosure_state=disclosure_state,
                as_of=row.as_of,
                provenance_ref=row.provenance_ref,
            )
        )
    return sorted(
        captured,
        key=lambda item: (item.asset_class, item.fund_name, item.provenance_ref),
    )


def ab2833_to_security_positions(
    disclosures: list[Ab2833AltDisclosure],
) -> list[SecurityPositionInput]:
    """Convert captured AB 2833 alts disclosures into security-position inputs.

    The vehicle's reported fair value becomes the position ``market_value_usd`` so
    alts feed the same reconciliation and coverage math as 13F/own-file holdings.
    """
    return [
        SecurityPositionInput(
            security_name=disclosure.fund_name,
            cusip=None,
            ticker=None,
            shares=None,
            market_value_usd=disclosure.reported_fair_value_usd,
            asset_class=disclosure.asset_class,
            source="ab2833",
            as_of=disclosure.as_of,
            provenance_ref=disclosure.provenance_ref,
            disclosure_state=disclosure.disclosure_state,
            manager_name=disclosure.manager_name,
            fund_name=disclosure.fund_name,
        )
        for disclosure in disclosures
    ]
